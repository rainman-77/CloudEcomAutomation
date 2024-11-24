import os
import openpyxl
from utilities.read_configurations import read_configuration as rc


def load_workbook_and_sheet(sheet_name):
    excel_path = rc("excel info", "path")
    excel_abs_file_path = os.path.abspath(excel_path)
    workbook = openpyxl.load_workbook(excel_abs_file_path)
    sheet = workbook[sheet_name]
    return workbook, sheet


def get_row_count(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.max_row


def get_column_count(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.max_column


def get_cell_data(sheet_name, row_num, column_num):
    _, sheet = load_workbook_and_sheet(sheet_name)
    return sheet.cell(row_num, column_num).value


def set_cell_data(sheet_name, row_num, column_num, data):
    excel_path = rc("excel info", "path")
    workbook, sheet = load_workbook_and_sheet(sheet_name)
    sheet.cell(row_num, column_num).value = data
    workbook.save(excel_path)


def get_all_excel_data(sheet_name):
    _, sheet = load_workbook_and_sheet(sheet_name)
    total_rows = sheet.max_row
    total_cols = sheet.max_column

    # Get the headers from the first row
    headers = [sheet.cell(1, c).value for c in range(1, total_cols + 1)]

    # Read data and map it to the headers
    final_list = []
    for r in range(2, total_rows + 1):  # Start from the second row for data
        row_dict = {headers[c - 1]: sheet.cell(r, c).value for c in range(1, total_cols + 1)}
        final_list.append(row_dict)

    return final_list
